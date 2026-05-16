"""
Reports API — generate business reports (PDF/XLSX) and list generation history.
Uses background tasks for actual report generation.
"""
import io
import os
import math
import random
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc
from app.db.postgres import get_db, async_session
from app.models.domain import ReportJobDB, OrderDB, ComplaintDB, UserDB
from app.models.schemas import ReportGenerateRequest
from app.api.auth import require_admin
from app.core.logger import logger

router = APIRouter()

# Directory to store generated reports
REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", "generated_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

MONTH_NAMES = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AUG", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC",
}


def _format_period(start_date: str, end_date: str) -> str:
    """Format date range into period string like 'JUL 01 - SEP 30'."""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        return f"{MONTH_NAMES[start.month]} {start.day:02d} - {MONTH_NAMES[end.month]} {end.day:02d}"
    except Exception:
        return f"{start_date} - {end_date}"


async def _generate_report_task(job_id: str, report_type: str, start_date: str, end_date: str, fmt: str):
    """Background task to generate the actual report file."""
    async with async_session() as db:
        try:
            # Find the job
            result = await db.execute(
                select(ReportJobDB).where(ReportJobDB.job_id == job_id)
            )
            job = result.scalar_one_or_none()
            if not job:
                logger.error(f"Report job {job_id} not found")
                return

            # Parse dates
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(end_date, "%Y-%m-%d")
            except ValueError:
                job.status = "failed"
                await db.commit()
                return

            # Fetch data based on report type
            if report_type == "SALES":
                rows = await db.execute(
                    select(OrderDB).where(
                        OrderDB.created_at >= start,
                        OrderDB.created_at <= end,
                        OrderDB.status.in_(("processing", "shipped", "completed")),
                    ).order_by(OrderDB.created_at)
                )
                data_rows = rows.scalars().all()
                report_data = [
                    {
                        "order_id": str(o.id)[:8],
                        "buyer": o.buyer_phone,
                        "amount": o.total_amount,
                        "status": o.status,
                        "date": o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
                    }
                    for o in data_rows
                ]
                title = f"Sales Report ({start_date} to {end_date})"
                columns = ["order_id", "buyer", "amount", "status", "date"]

            elif report_type == "COMPLAINTS":
                rows = await db.execute(
                    select(ComplaintDB).where(
                        ComplaintDB.created_at >= start,
                        ComplaintDB.created_at <= end,
                    ).order_by(ComplaintDB.created_at)
                )
                data_rows = rows.scalars().all()
                report_data = [
                    {
                        "complaint_id": str(c.id)[:8],
                        "buyer": c.buyer_phone,
                        "description": (c.description or "")[:100],
                        "status": c.status,
                        "date": c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
                    }
                    for c in data_rows
                ]
                title = f"Complaints Report ({start_date} to {end_date})"
                columns = ["complaint_id", "buyer", "description", "status", "date"]

            elif report_type == "REFUNDS":
                rows = await db.execute(
                    select(ComplaintDB).where(
                        ComplaintDB.created_at >= start,
                        ComplaintDB.created_at <= end,
                        ComplaintDB.status == "approved",
                    ).order_by(ComplaintDB.created_at)
                )
                data_rows = rows.scalars().all()
                report_data = [
                    {
                        "complaint_id": str(c.id)[:8],
                        "buyer": c.buyer_phone,
                        "description": (c.description or "")[:100],
                        "date": c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
                    }
                    for c in data_rows
                ]
                title = f"Refunds Report ({start_date} to {end_date})"
                columns = ["complaint_id", "buyer", "description", "date"]

            elif report_type == "RECEIPTS":
                rows = await db.execute(
                    select(OrderDB).where(
                        OrderDB.created_at >= start,
                        OrderDB.created_at <= end,
                        OrderDB.status == "completed",
                    ).order_by(OrderDB.created_at)
                )
                data_rows = rows.scalars().all()
                report_data = [
                    {
                        "order_id": str(o.id)[:8],
                        "buyer": o.buyer_phone,
                        "amount": o.total_amount,
                        "date": o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
                    }
                    for o in data_rows
                ]
                title = f"Receipts Report ({start_date} to {end_date})"
                columns = ["order_id", "buyer", "amount", "date"]
            else:
                job.status = "failed"
                await db.commit()
                return

            # Generate the file
            report_id = job.report_id
            filename = f"{report_type}_{start_date}_to_{end_date}_{report_id}"

            if fmt == "XLSX":
                filepath = os.path.join(REPORTS_DIR, f"{filename}.xlsx")
                _generate_xlsx(filepath, title, columns, report_data)
            else:
                filepath = os.path.join(REPORTS_DIR, f"{filename}.pdf")
                _generate_pdf(filepath, title, columns, report_data)

            # Update job status
            job.status = "completed"
            job.download_url = f"/api/v1/reports/download/{report_id}"
            await db.commit()
            logger.info(f"Report {report_id} generated successfully: {filepath}")

        except Exception as e:
            logger.error(f"Failed to generate report {job_id}: {e}")
            try:
                result = await db.execute(
                    select(ReportJobDB).where(ReportJobDB.job_id == job_id)
                )
                job = result.scalar_one_or_none()
                if job:
                    job.status = "failed"
                    await db.commit()
            except Exception:
                pass


def _generate_pdf(filepath: str, title: str, columns: list, data: list):
    """Generate a basic PDF report using reportlab if available, otherwise plain text."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 20))

        if data:
            table_data = [columns]
            for row in data:
                table_data.append([str(row.get(c, "")) for c in columns])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No data found for the specified period.", styles['Normal']))

        doc.build(elements)

    except ImportError:
        # Fallback: write a simple text file with .pdf extension
        with open(filepath, 'w') as f:
            f.write(f"{title}\n{'='*60}\n\n")
            if data:
                f.write("\t".join(columns) + "\n")
                f.write("-" * 60 + "\n")
                for row in data:
                    f.write("\t".join(str(row.get(c, "")) for c in columns) + "\n")
            else:
                f.write("No data found for the specified period.\n")


def _generate_xlsx(filepath: str, title: str, columns: list, data: list):
    """Generate an XLSX report using openpyxl if available, otherwise CSV fallback."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True)

        # Header row
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name.upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(data, 4):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        # Auto-fit columns
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 20

        wb.save(filepath)

    except ImportError:
        # Fallback: write CSV
        import csv
        csv_path = filepath.replace(".xlsx", ".csv")
        with open(csv_path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)
        # Rename to xlsx for consistency
        os.rename(csv_path, filepath)


@router.post("/generate")
async def generate_report(
    payload: ReportGenerateRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(require_admin),
):
    """Submits a job to generate a business report."""
    # Generate IDs
    seq = random.randint(100, 999)
    job_id = f"JOB-{seq}"
    report_id = f"REP-{seq}"
    period = _format_period(payload.start_date, payload.end_date)
    name = f"{payload.report_type.capitalize()}_Report_{payload.start_date}_to_{payload.end_date}"

    # Create job record
    job = ReportJobDB(
        job_id=job_id,
        report_id=report_id,
        report_type=payload.report_type,
        name=name,
        start_date=payload.start_date,
        end_date=payload.end_date,
        period=period,
        format=payload.format,
        status="processing",
    )
    db.add(job)
    await db.commit()

    # Enqueue background task
    background_tasks.add_task(
        _generate_report_task, job_id, payload.report_type, payload.start_date, payload.end_date, payload.format
    )

    return {"status": "processing", "job_id": job_id, "full_id": str(job.id)}


@router.get("/history")
async def report_history(
    db: AsyncSession = Depends(get_db),
    current_user: UserDB = Depends(require_admin),
):
    """Fetches the last generated reports available for download."""
    stmt = select(ReportJobDB).order_by(desc(ReportJobDB.created_at)).limit(50)
    result = await db.execute(stmt)
    jobs = result.scalars().all()

    items = []
    for job in jobs:
        items.append({
            "report_id": job.report_id or "",
            "full_id": job.id,
            "name": job.name or "",
            "generation_type": job.generation_type or "Manual Export",
            "created_at": job.created_at.isoformat() + "Z" if job.created_at else "",
            "period": job.period or "",
            "format": job.format or "PDF",
            "status": job.status,
            "download_url": job.download_url or "",
        })

    return items


@router.get("/download/{report_id}")
async def download_report(
    report_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Download a generated report file."""
    result = await db.execute(
        select(ReportJobDB).where(ReportJobDB.report_id == report_id)
    )
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Report not found")

    if job.status != "completed":
        raise HTTPException(status_code=400, detail=f"Report is still {job.status}")

    # Find the file
    ext = ".xlsx" if job.format == "XLSX" else ".pdf"
    filename = f"{job.report_type}_{job.start_date}_to_{job.end_date}_{report_id}{ext}"
    filepath = os.path.join(REPORTS_DIR, filename)

    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Report file not found on disk")

    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext == ".xlsx" else "application/pdf"
    return FileResponse(filepath, filename=filename, media_type=media_type)
